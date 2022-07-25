from json import load
from openpyxl import workbook, load_workbook

wb = load_workbook('Pyxl/Trab.xlsx')
ws = wb.active
print(ws['A1'].value)

wb.save('Pyxl/Trab.xlsx')



